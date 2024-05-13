import tkinter as tk
from tkinter import filedialog
import openpyxl

NotaCorte = 7.0
PesoQualidade = 6.0
PesoQnt = 2.5
PesoPontualidade = 1.5
Pontualidade = 0
Quantidade = 0
PontoInspecao = 0
Agrupamento = []
consolidatedData = {}


def process_excel(file):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        valueA, valueB, valueC, valueD, valueE, valueF, valueG = row[:7]

        valueF = float(str(valueF).replace('.', '').replace(',', '.'))
        valueG = float(str(valueG).replace('.', '').replace(',', '.'))
        valueA = valueA.upper()
        valueC = valueC.upper()  # Convert razão social to uppercase

        if valueA == "INSPEÇÃO APROVADA":
            PontoInspecao = 1
        elif valueA == "INSPEÇÃO REPROVADA":
            PontoInspecao = 0

        else:
            print(f"Código: {valueB}, Razão Social: {valueC}")
            PontoInspecao = 0
          
            

        Pontualidade = 1 if valueE <= valueD else 0
        Quantidade = 1 if valueG >= valueF else 0

        if valueB not in Agrupamento:
            Agrupamento.append(valueB)

        entry = {
            "Codigo": valueB,
            "Razão Social": valueC,
            "Inspeção": PontoInspecao,
            "Pontualidade": Pontualidade,
            "Quantidade": Quantidade
        }

        
        if entry["Codigo"] + entry["Razão Social"] not in consolidatedData:
            consolidatedData[entry["Codigo"] + entry["Razão Social"]] = {
                "Codigo": entry["Codigo"],
                "Razão Social": entry["Razão Social"],
                "Inspeção": 0,
                "Pontualidade": 0,
                "Quantidade": 0,
                "Count": 0
            }

        consolidatedData[entry["Codigo"] + entry["Razão Social"]]["Inspeção"] += entry["Inspeção"]
        consolidatedData[entry["Codigo"] + entry["Razão Social"]]["Pontualidade"] += entry["Pontualidade"]
        consolidatedData[entry["Codigo"] + entry["Razão Social"]]["Quantidade"] += entry["Quantidade"]
        consolidatedData[entry["Codigo"] + entry["Razão Social"]]["Count"] += 1

    agrupapador()


def agrupapador():
    global consolidatedData
    for entry_key in consolidatedData:
        entry = consolidatedData[entry_key]
        totalInspecao = entry["Inspeção"] / entry["Count"]
        totalPontualidade = entry["Pontualidade"] / entry["Count"]
        totalQuantidade = entry["Quantidade"] / entry["Count"]

        entry["Inspeção"] = round(totalInspecao, 2)
        entry["Pontualidade"] = round(totalPontualidade, 2)
        entry["Quantidade"] = round(totalQuantidade, 2)

        media = round((totalInspecao * PesoQualidade) + (totalPontualidade * PesoPontualidade) + (totalQuantidade * PesoQnt), 2)
        entry["IMF"] = media
        entry["Resultado"] = "Aprovado" if media >= NotaCorte else "Reprovado"

        del entry["Count"]

    consolidatedData = sorted(consolidatedData.values(), key=lambda x: x["Razão Social"].upper())


def select_file_and_generate_excel():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        process_excel(file_path)
        save_excel_file()


def save_excel_file():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.asksaveasfilename(title="Save Excel file", defaultextension=".xlsx",
                                              filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        # Write headers
        sheet.append(["Código", "Razão Social", "Média de Inspeção", "Média de Pontualidade",
                      "Média de Quantidade", "IMF", "Resultado"])

        # Write data
        for entry in consolidatedData:
            sheet.append([entry["Codigo"], entry["Razão Social"], entry["Inspeção"], entry["Pontualidade"],
                          entry["Quantidade"], entry["IMF"], entry["Resultado"]])

        workbook.save(file_path)
        print("Excel file saved successfully!")


if __name__ == "__main__":
    select_file_and_generate_excel()
